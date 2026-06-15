"""Excel I/O module for reading questions and writing results."""
import os
from pathlib import Path
from typing import List, Optional
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from logger import pipeline_logger
from models.schemas import Question, TestRow
from config import settings

def read_questions(file_path: str = settings.input_excel, sheet_name: str = settings.input_sheet) -> List[Question]:
    """Reads questions from an input Excel file."""
    if not os.path.exists(file_path):
        pipeline_logger.log_simple_error(f"Input file not found: {file_path}")
        return []

    wb = openpyxl.load_workbook(file_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        pipeline_logger.log_simple_error(f"Sheet {sheet_name} not found in {file_path}")
        return []
        
    sheet = wb[sheet_name]
    
    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header_row:
        pipeline_logger.log_simple_error("No header row found in input Excel")
        return []
        
    try:
        q_idx = header_row.index(settings.question_column)
    except ValueError:
        pipeline_logger.log_simple_error(f"Column '{settings.question_column}' not found in header: {header_row}")
        return []
        
    try:
        id_idx = header_row.index(settings.id_column)
    except ValueError:
        id_idx = -1
        pipeline_logger.log_simple_info(f"ID column '{settings.id_column}' not found, generating IDs automatically")
        
    questions = []
    for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if not row or len(row) <= q_idx or not row[q_idx]:
            continue
            
        q_text = str(row[q_idx]).strip()
        if not q_text:
            continue
            
        q_id = str(row[id_idx]) if id_idx >= 0 and len(row) > id_idx and row[id_idx] is not None else str(i)
        questions.append(Question(id=q_id, text=q_text))
        
    return questions

def write_results(results: List[TestRow], output_path: Optional[Path] = None) -> None:
    """Writes the test results to an output Excel file with styling.
    
    If output_path is provided (used by CheckpointManager), writes to that exact path.
    Otherwise creates a timestamped file in settings.output_dir.
    """
    if not results:
        pipeline_logger.log_warning("No results to write")
        return
        
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Results"
    
    headers = [
        "id", "question", "model", "answer", "answer_raw", 
        "chunks_used", "avg_chunk_score", "quality_score", "quality_reason", "quality_rank",
        "embed_ms", "retrieve_ms", "llm_ms", "total_ms", 
        "prompt_tokens", "completion_tokens", "tokens_per_sec", "error"
    ]
    sheet.append(headers)
    
    # Write data
    for row_data in results:
        q = row_data.question
        r = row_data.result
        
        row = [
            q.id,
            q.text,
            r.model,
            r.answer,
            r.answer_raw,
            r.chunks_used,
            round(r.avg_score, 4),
            r.quality_score,
            r.quality_reason,
            r.quality_rank,
            round(r.embed_ms, 2),
            round(r.retrieve_ms, 2),
            round(r.llm_ms, 2),
            round(r.total_ms, 2),
            round(r.tokens_per_sec, 2),
            r.prompt_tokens,
            r.completion_tokens,
            r.error
        ]
        sheet.append(row)
        
    # Styling
    header_fill = PatternFill(start_color="DDEEFF", end_color="DDEEFF", fill_type="solid")
    header_font = Font(bold=True)
    even_row_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
    
    # Format header
    for cell in sheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        
    # Auto-filter and freeze top row
    sheet.auto_filter.ref = sheet.dimensions
    sheet.freeze_panes = "A2"
    
    # Column widths
    column_widths = {
        "A": 8,   # id
        "B": 50,  # question
        "C": 20,  # model
        "D": 80,  # answer
        "E": 80,  # answer_raw
        "F": 12,  # chunks_used
        "G": 12,  # avg_score
        "H": 12,  # embed_ms
        "I": 12,  # retrieve_ms
        "J": 12,  # llm_ms
        "K": 12,  # total_ms
        "L": 14,  # tokens_per_sec
        "M": 14,  # prompt_tokens
        "N": 16,  # completion_tokens
        "O": 40   # error
    }
    
    for col_letter, width in column_widths.items():
        if col_letter in sheet.column_dimensions:
            pass # just a check
        sheet.column_dimensions[col_letter].width = width
        
    # Wrap text and alternate shading
    wrap_alignment = Alignment(wrap_text=True, vertical="top")
    
    for row_idx, row in enumerate(sheet.iter_rows(min_row=2), start=2):
        row[1].alignment = wrap_alignment  # question
        row[3].alignment = wrap_alignment  # answer
        row[4].alignment = wrap_alignment  # answer_raw
        
        if row_idx % 2 == 0:
            for cell in row:
                cell.fill = even_row_fill
                
    # Save to exact path (checkpoint) or timestamped path in output_dir
    if output_path is not None:
        save_path = Path(output_path)
    else:
        out_dir = Path(settings.output_dir)
        out_dir.mkdir(parents=True, exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        save_path = out_dir / f"results_{timestamp}.xlsx"

    wb.save(str(save_path))
    pipeline_logger.log_simple_info(f"Results saved to {save_path}")
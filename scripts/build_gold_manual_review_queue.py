from __future__ import annotations

import argparse
import csv
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook, load_workbook


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Build manual review queue for unresolved gold citations.")
    parser.add_argument("--source-xlsx", default="tests/benchmarks/Полный бенчмарк-3.xlsx")
    parser.add_argument("--normalized-xlsx", default="tests/benchmarks/Полный бенчмарк-3.normalized.xlsx")
    parser.add_argument("--audit-csv", default="tests/benchmarks/gold_citations_audit.normalized.csv")
    parser.add_argument("--sheet", default="Benchmark")
    parser.add_argument("--output", default="tests/benchmarks/manual_review_queue.xlsx")
    return parser.parse_args()


def clean_text(value: object) -> str:
    if value is None:
        return ""
    return " ".join(str(value).replace("\xa0", " ").split()).strip()


def load_rows(xlsx_path: Path, sheet_name: str) -> dict[int, dict[str, str]]:
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb[sheet_name]
    headers = [clean_text(v) for v in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    rows: dict[int, dict[str, str]] = {}
    for row_number, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        payload = {headers[idx]: row[idx] for idx in range(min(len(headers), len(row)))}
        rows[row_number] = {
            "query_id": clean_text(payload.get("query_id")),
            "query": clean_text(payload.get("query")),
            "answer": clean_text(payload.get("Ответ нашего RAG")),
            "gold_citations": clean_text(payload.get("gold_citations")),
        }
    return rows


def main() -> None:
    args = parse_args()
    source_rows = load_rows(Path(args.source_xlsx), args.sheet)
    normalized_rows = load_rows(Path(args.normalized_xlsx), args.sheet)

    grouped: dict[tuple[int, str], dict[str, object]] = defaultdict(lambda: {"issues": [], "citations": []})
    with open(args.audit_csv, encoding="utf-8") as handle:
        reader = csv.DictReader(handle)
        for item in reader:
            row_number = int(item["row_number"])
            key = (row_number, item["query_id"])
            grouped[key]["issues"].append(item["issue"])
            grouped[key]["citations"].append(item["citation"])

    wb = Workbook()
    ws = wb.active
    ws.title = "manual_review"
    ws.append(
        [
            "row_number",
            "query_id",
            "issues",
            "flagged_citations",
            "original_gold_citations",
            "normalized_gold_citations",
            "query",
            "answer_excerpt",
        ]
    )

    for (row_number, query_id), payload in sorted(grouped.items()):
        source = source_rows.get(row_number, {})
        normalized = normalized_rows.get(row_number, {})
        issues = "; ".join(sorted(set(str(v) for v in payload["issues"])))
        citations = "; ".join(dict.fromkeys(str(v) for v in payload["citations"]))
        answer_excerpt = clean_text(source.get("answer", ""))[:1200]
        ws.append(
            [
                row_number,
                query_id,
                issues,
                citations,
                source.get("gold_citations", ""),
                normalized.get("gold_citations", ""),
                source.get("query", ""),
                answer_excerpt,
            ]
        )

    output_path = Path(args.output)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f"Saved manual review queue: {output_path}")
    print(f"Rows queued: {ws.max_row - 1}")


if __name__ == "__main__":
    main()

from __future__ import annotations

import argparse
from pathlib import Path

from openpyxl import Workbook, load_workbook


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Remove rows listed in a manual review queue from a benchmark workbook."
    )
    parser.add_argument(
        "--input",
        default="tests/benchmarks/Полный бенчмарк-3.reviewed9.xlsx",
        help="Reviewed benchmark workbook.",
    )
    parser.add_argument(
        "--queue",
        default="tests/benchmarks/manual_review_queue.reviewed9.xlsx",
        help="Manual review queue workbook with row_number column.",
    )
    parser.add_argument(
        "--sheet",
        default="Benchmark",
        help="Worksheet name in the benchmark workbook.",
    )
    parser.add_argument(
        "--queue-sheet",
        default="manual_review",
        help="Worksheet name in the manual review queue workbook.",
    )
    parser.add_argument(
        "--output",
        default="tests/benchmarks/kz_benchmark_gold_final.filtered.xlsx",
        help="Filtered output workbook.",
    )
    return parser.parse_args()


def clean_text(value: object) -> str:
    if value is None:
        return ""
    return " ".join(str(value).replace("\xa0", " ").split()).strip()


def load_excluded_rows(queue_path: Path, sheet_name: str) -> set[int]:
    wb = load_workbook(queue_path, read_only=True, data_only=True)
    ws = wb[sheet_name]
    headers = [clean_text(v) for v in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    row_index = headers.index("row_number")

    excluded: set[int] = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[row_index] is None:
            continue
        excluded.add(int(row[row_index]))
    return excluded


def main() -> None:
    args = parse_args()
    excluded = load_excluded_rows(Path(args.queue), args.queue_sheet)

    source_wb = load_workbook(args.input, read_only=True, data_only=False)
    source_ws = source_wb[args.sheet]

    out_wb = Workbook()
    out_ws = out_wb.active
    out_ws.title = args.sheet

    kept = 0
    removed = 0
    for row_number, row in enumerate(source_ws.iter_rows(values_only=True), start=1):
        if row_number == 1:
            out_ws.append(list(row))
            continue
        if row_number in excluded:
            removed += 1
            continue
        out_ws.append(list(row))
        kept += 1

    output_path = Path(args.output)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    out_wb.save(output_path)
    print(f"Saved filtered benchmark: {output_path}")
    print(f"Removed rows: {removed}")
    print(f"Kept data rows: {kept}")


if __name__ == "__main__":
    main()

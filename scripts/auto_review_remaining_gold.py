from __future__ import annotations

import argparse
import re
from collections import Counter, OrderedDict
from pathlib import Path

from openpyxl import load_workbook


ACT_PATTERNS: list[tuple[re.Pattern[str], str]] = [
    (re.compile(r"закон[а-я\s]*(?:республики казахстан|рк)?\s*[«\"]об исполнительном производстве и статусе судебных исполнителей[»\"]?", re.IGNORECASE), "Закон РК «Об исполнительном производстве и статусе судебных исполнителей»"),
    (re.compile(r"закон[а-я\s]*(?:республики казахстан|рк)?\s*[«\"]о защите прав потребителей[»\"]?", re.IGNORECASE), "Закон РК «О защите прав потребителей»"),
    (re.compile(r"закон[а-я\s]*(?:республики казахстан|рк)?\s*[«\"]о жилищных отношениях[»\"]?", re.IGNORECASE), "Закон РК «О жилищных отношениях»"),
    (re.compile(r"закон[а-я\s]*(?:республики казахстан|рк)?\s*[«\"]об образовании[»\"]?", re.IGNORECASE), "Закон РК «Об образовании»"),
    (re.compile(r"закон[а-я\s]*(?:республики казахстан|рк)?\s*[«\"]о воинской службе и статусе военнослужащих[»\"]?", re.IGNORECASE), "Закон РК «О воинской службе и статусе военнослужащих»"),
    (re.compile(r"закон[а-я\s]*(?:республики казахстан|рк)?\s*[«\"]о специальных государственных органах республики казахстан[»\"]?", re.IGNORECASE), "Закон РК «О специальных государственных органах Республики Казахстан»"),
    (re.compile(r"закон[а-я\s]*(?:республики казахстан|рк)?\s*[«\"]о государственной регистрации прав на недвижимое имущество[»\"]?", re.IGNORECASE), "Закон РК «О государственной регистрации прав на недвижимое имущество»"),
    (re.compile(r"закон[а-я\s]*(?:республики казахстан|рк)?\s*[«\"]об обязательном страховании гражданско-правовой ответственности владельцев транспортных средств[»\"]?", re.IGNORECASE), "Закон РК «Об обязательном страховании гражданско-правовой ответственности владельцев транспортных средств»"),
    (re.compile(r"закон[а-я\s]*(?:республики казахстан|рк)?\s*[«\"]о правах ребенка(?: в республике казахстан)?[»\"]?", re.IGNORECASE), "Закон РК «О правах ребенка в Республике Казахстан»"),
    (re.compile(r"закон[а-я\s]*(?:республики казахстан|рк)?\s*[«\"]о социальной защите лиц с инвалидностью в республике казахстан[»\"]?", re.IGNORECASE), "Закон РК «О социальной защите лиц с инвалидностью в Республике Казахстан»"),
    (re.compile(r"закон[а-я\s]*(?:республики казахстан|рк)?\s*[«\"]о банках и банковской деятельности в республике казахстан[»\"]?", re.IGNORECASE), "Закон РК «О банках и банковской деятельности в Республике Казахстан»"),
    (re.compile(r"закон[а-я\s]*(?:республики казахстан|рк)?\s*[«\"]о государственном регулировании, контроле и надзоре финансового рынка и финансовых организаций[»\"]?", re.IGNORECASE), "Закон РК «О государственном регулировании, контроле и надзоре финансового рынка и финансовых организаций»"),
    (re.compile(r"закон[а-я\s]*(?:республики казахстан|рк)?\s*[«\"]о восстановлении платежеспособности и банкротстве граждан(?: республики казахстан)?[»\"]?", re.IGNORECASE), "Закон РК «О восстановлении платежеспособности и банкротстве граждан Республики Казахстан»"),
    (re.compile(r"закон[а-я\s]*(?:республики казахстан|рк)?\s*[«\"]об авторском праве и смежных правах[»\"]?", re.IGNORECASE), "Закон РК «Об авторском праве и смежных правах»"),
    (re.compile(r"гражданск[а-я\s]+кодекс[а-я\s]*(?:республики казахстан|рк)", re.IGNORECASE), "ГК РК"),
    (re.compile(r"\bгк рк\b", re.IGNORECASE), "ГК РК"),
    (re.compile(r"гражданск[а-я\s]+процессуальн[а-я\s]+кодекс[а-я\s]*(?:республики казахстан|рк)", re.IGNORECASE), "ГПК РК"),
    (re.compile(r"\bгпк рк\b", re.IGNORECASE), "ГПК РК"),
    (re.compile(r"административн[а-я\s]+процедурно-процессуальн[а-я\s]+кодекс[а-я\s]*(?:республики казахстан|рк)", re.IGNORECASE), "АППК РК"),
    (re.compile(r"\bаппк\b|\bаппк рк\b", re.IGNORECASE), "АППК РК"),
    (re.compile(r"трудов[а-я\s]+кодекс[а-я\s]*(?:республики казахстан|рк)", re.IGNORECASE), "ТК РК"),
    (re.compile(r"\bтк рк\b", re.IGNORECASE), "ТК РК"),
    (re.compile(r"уголовн[а-я\s]+кодекс[а-я\s]*(?:республики казахстан|рк)", re.IGNORECASE), "УК РК"),
    (re.compile(r"\bук рк\b", re.IGNORECASE), "УК РК"),
    (re.compile(r"кодекс[а-я\s]+о браке \(супружестве\) и семье", re.IGNORECASE), "Кодекс РК «О браке (супружестве) и семье»"),
    (re.compile(r"семейн[а-я\s]+кодекс", re.IGNORECASE), "Кодекс РК «О браке (супружестве) и семье»"),
    (re.compile(r"земельн[а-я\s]+кодекс[а-я\s]*(?:республики казахстан|рк)", re.IGNORECASE), "ЗК РК"),
    (re.compile(r"\bзк рк\b", re.IGNORECASE), "ЗК РК"),
    (re.compile(r"конституци[а-я\s]*(?:республики казахстан|рк)", re.IGNORECASE), "Конституция РК"),
]

STRUCT_UNIT_ONLY_RE = re.compile(r"^(подп\.\s*[^;]+|п\.\s*\d+\)?|ч\.\s*\d+)$", re.IGNORECASE)
ARTICLE_RE = re.compile(r"^(?:(?:пп?\.\s*[^;]+?\s*)?(?:п\.\s*[\d-]+\s*)?(?:ч\.\s*[\d-]+\s*)?)?ст\.\s*[\d-]+$", re.IGNORECASE)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Conservative auto-review for remaining unresolved gold citations.")
    parser.add_argument("--input", default="tests/benchmarks/Полный бенчмарк-3.reviewed.xlsx")
    parser.add_argument("--output", default="tests/benchmarks/Полный бенчмарк-3.reviewed2.xlsx")
    parser.add_argument("--queue", default="tests/benchmarks/manual_review_queue.reviewed.xlsx")
    parser.add_argument("--sheet", default="Benchmark")
    return parser.parse_args()


def clean_text(value: object) -> str:
    if value is None:
        return ""
    return " ".join(str(value).replace("\xa0", " ").split()).strip()


def extract_acts(text: str) -> list[str]:
    acts: list[str] = []
    for pattern, act in ACT_PATTERNS:
        for _ in pattern.finditer(text):
            acts.append(act)
    return acts


def get_dominant_act(text: str) -> str:
    acts = extract_acts(text)
    if not acts:
        return ""
    counts = Counter(acts)
    act, n = counts.most_common(1)[0]
    if len(counts) == 1:
        return act
    if n >= 2 and n > counts.most_common(2)[1][1]:
        return act
    return ""


def normalize_flagged(flagged: str, dominant_act: str) -> str:
    items = [clean_text(part) for part in flagged.split(";") if clean_text(part)]
    out: list[str] = []
    previous_article = ""
    for item in items:
        if ARTICLE_RE.match(item):
            out.append(f"{item} {dominant_act}")
            article_match = re.search(r"ст\.\s*([\d-]+)", item, re.IGNORECASE)
            previous_article = article_match.group(1) if article_match else previous_article
            continue
        if STRUCT_UNIT_ONLY_RE.match(item) and previous_article:
            out.append(f"{item} ст. {previous_article} {dominant_act}")
            continue
        out.append(item)
    return "; ".join(OrderedDict.fromkeys(out))


def main() -> None:
    args = parse_args()
    queue_wb = load_workbook(args.queue, read_only=True, data_only=True)
    queue_ws = queue_wb["manual_review"]
    candidates: dict[int, tuple[str, str]] = {}

    for row in queue_ws.iter_rows(min_row=2, values_only=True):
        row_number, _, issues, flagged, _, normalized, query, excerpt = row
        text = f"{clean_text(query)}\n{clean_text(excerpt)}"
        dominant = get_dominant_act(text)
        if not dominant:
            continue
        flagged_text = clean_text(flagged)
        if not flagged_text:
            continue
        if "missing_act_name" not in clean_text(issues) and "structural_unit_without_article_or_act" not in clean_text(issues):
            continue
        candidate = normalize_flagged(flagged_text, dominant)
        if candidate and candidate != clean_text(normalized):
            candidates[int(row_number)] = (flagged_text, candidate)

    wb = load_workbook(args.input)
    ws = wb[args.sheet]
    headers = [clean_text(v) for v in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    gold_col = headers.index("gold_citations") + 1
    applied = 0
    for row_number, payload in candidates.items():
        flagged_text, candidate = payload
        current = clean_text(ws.cell(row_number, gold_col).value)
        if not current:
            continue
        parts = [clean_text(part) for part in current.split(";") if clean_text(part)]
        flagged_parts = {clean_text(part) for part in flagged_text.split(";") if clean_text(part)}
        replacement_parts = [clean_text(part) for part in candidate.split(";") if clean_text(part)]
        kept = [part for part in parts if part not in flagged_parts]
        merged = list(OrderedDict.fromkeys([*kept, *replacement_parts]))
        updated = "; ".join(merged)
        if updated != current:
            ws.cell(row_number, gold_col).value = updated
            applied += 1

    output_path = Path(args.output)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f"Saved auto-reviewed workbook: {output_path}")
    print(f"Rows updated: {applied}")


if __name__ == "__main__":
    main()

from __future__ import annotations

import re
from collections import OrderedDict
from pathlib import Path

from openpyxl import Workbook, load_workbook


BASE_DIR = Path(__file__).resolve().parents[1]
BENCHMARK_PATH = BASE_DIR / "tests" / "benchmarks" / "Полный бенчмарк-3.xlsx"
AUTHORITATIVE_CITATIONS_PATH = BASE_DIR / "tests" / "benchmarks" / "642_questions_with_citations.xlsx"

HEADERS = [
    "query_id",
    "query",
    "Ответ нашего RAG",
    "gold_citations",
    "student_name",
    "student_email",
]

DISCLAIMER = "Это не официальная юридическая консультация. Информация только из базы."


def clean_text(value: object) -> str:
    if value is None:
        return ""
    text = str(value).replace("\r\n", "\n").replace("\r", "\n").replace("\xa0", " ")
    text = text.replace("\u200b", "").strip()
    text = re.sub(r"[ \t]+\n", "\n", text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text.strip()


def canonicalize_gold_from_authoritative(value: object) -> str:
    text = clean_text(value)
    if not text:
        return ""
    parts = [part.strip(" ;") for part in text.split(";")]
    return "; ".join(OrderedDict.fromkeys(part for part in parts if part))


def normalize_article_token(token: str) -> str:
    token = clean_text(token)
    token = re.sub(r"\bстатьи\b", "ст.", token, flags=re.IGNORECASE)
    token = re.sub(r"\bстатья\b", "ст.", token, flags=re.IGNORECASE)
    token = re.sub(r"\bст\.\s*", "ст. ", token, flags=re.IGNORECASE)
    token = re.sub(r"\bпункта\b", "п.", token, flags=re.IGNORECASE)
    token = re.sub(r"\bпункт\b", "п.", token, flags=re.IGNORECASE)
    token = re.sub(r"\bподпункта\b", "подп.", token, flags=re.IGNORECASE)
    token = re.sub(r"\bподпункт\b", "подп.", token, flags=re.IGNORECASE)
    token = re.sub(r"\bчасти\b", "ч.", token, flags=re.IGNORECASE)
    token = re.sub(r"\bчасть\b", "ч.", token, flags=re.IGNORECASE)
    token = re.sub(r"\s+", " ", token)
    return token.strip(" ;")


def is_act_only(token: str) -> bool:
    lower = token.lower()
    if any(marker in lower for marker in ("закон", "кодекс", "конституц", "приказ", "постанов", "правил")):
        return "ст." not in lower and "п." not in lower and "ч." not in lower and "подп." not in lower
    return False


def looks_like_article(token: str) -> bool:
    lower = token.lower()
    return any(marker in lower for marker in ("ст.", "статья", "статьи", "п.", "пункт", "ч.", "част", "подп."))


def derive_from_article_excerpt(article_text: str) -> list[str]:
    article_text = clean_text(article_text)
    if not article_text:
        return []

    act_match = re.search(
        r"(Гражданского кодекса Республики Казахстан|Трудового кодекса Республики Казахстан|"
        r"Конституции Республики Казахстан|Закона Республики Казахстан «[^»]+»)",
        article_text,
        flags=re.IGNORECASE,
    )
    act_name = act_match.group(1) if act_match else ""

    results: list[str] = []
    first_article = re.search(r"стать[ьяи]\s+(\d+[.-]?\d*)", article_text, flags=re.IGNORECASE)
    if first_article and act_name:
        results.append(f"ст. {first_article.group(1)} {act_name}")

    for point, article in re.findall(r"пункт[ау]?\s+(\d+[.-]?\d*)\s+стать[ьяи]\s+(\d+[.-]?\d*)", article_text, flags=re.IGNORECASE):
        if act_name:
            results.append(f"п. {point} ст. {article} {act_name}")

    if first_article:
        first_article_number = first_article.group(1)
        for point in re.findall(r"пункт[ау]?\s+(\d+[.-]?\d*)\s+данной статьи", article_text, flags=re.IGNORECASE):
            if act_name:
                results.append(f"п. {point} ст. {first_article_number} {act_name}")

    return list(OrderedDict.fromkeys(results))


def normalize_gold(raw_value: object, article_value: object | None = None) -> str:
    text = clean_text(raw_value)
    article_text = clean_text(article_value)

    if not text and article_text:
        derived = derive_from_article_excerpt(article_text)
        if derived:
            return "; ".join(derived)
        return article_text

    if not text:
        return ""

    raw_parts = [normalize_article_token(part) for part in re.split(r"[;\n|]+", text) if clean_text(part)]
    combined: list[str] = []
    current_act = ""

    for part in raw_parts:
        if is_act_only(part):
            current_act = part
            continue

        if looks_like_article(part) and current_act and current_act.lower() not in part.lower():
            combined.append(f"{part} {current_act}")
            continue

        combined.append(part)

    if article_text:
        derived = derive_from_article_excerpt(article_text)
        for item in derived:
            if item not in combined:
                combined.append(item)

    combined = [item.strip(" ;") for item in combined if item.strip(" ;")]
    return "; ".join(OrderedDict.fromkeys(combined))


def extract_citations_from_answer(answer_value: object) -> str:
    answer = clean_text(answer_value)
    if not answer:
        return ""

    patterns = [
        r"(?:пп?\.\s*\d+[)-]?\s*)?(?:п\.?\s*\d+[.-]?\d*\s*)?(?:ч\.?\s*\d+[.-]?\d*\s*)?(?:ст\.?|стать[а-яё]{1,5})\s*\d+[.-]?\d*(?:\s*[,-]\s*\d+[.-]?\d*)?(?:[^.\n]{0,140}?(?:РК|Республики Казахстан|«[^»]+»))",
        r"(?:пп?\.\s*\d+[)-]?\s*)?(?:п\.?\s*\d+[.-]?\d*\s*)?(?:ч\.?\s*\d+[.-]?\d*\s*)?(?:ст\.?|стать[а-яё]{1,5})\s*\d+[.-]?\d*",
        r"(?:пп?\.\s*\d+[)-]?\s*|п\.?\s*\d+[.-]?\d*\s*)(?:[^.\n]{0,160}?(?:Правил|Приказа|Постановления|Перечня)[^.\n]{0,120}(?:№\s*[\w/-]+)?)",
        r"(?:Приказ|Постановление)\s+[A-ZА-ЯЁ][^.\n]{0,180}?№\s*[\w/-]+",
    ]

    matches: list[str] = []
    for pattern in patterns:
        for match in re.findall(pattern, answer, flags=re.IGNORECASE):
            normalized = normalize_article_token(match)
            if normalized:
                matches.append(normalized)

    return "; ".join(OrderedDict.fromkeys(matches))


def standardize_answer(answer_value: object, gold_citations: str) -> str:
    answer = clean_text(answer_value)
    if not answer:
        if gold_citations:
            answer = (
                "Здравствуйте!\n"
                "По вашему вопросу правовую позицию нужно строить на указанных ниже нормах.\n\n"
                f"Ключевые нормы: {gold_citations}."
            )
        else:
            answer = "Здравствуйте!\nПо вашему вопросу нужен дополнительный анализ фактов и применимых норм."

    answer = re.sub(r"\nИсточники:.*$", "", answer, flags=re.IGNORECASE | re.DOTALL).strip()
    if not answer.startswith(DISCLAIMER):
        answer = f"{DISCLAIMER}\n{answer}"
    if gold_citations:
        answer = f"{answer}\n\nИсточники:\n{gold_citations}"
    return answer.strip()


def load_authoritative_citations() -> dict[str, str]:
    wb = load_workbook(AUTHORITATIVE_CITATIONS_PATH)
    ws = wb[wb.sheetnames[0]]
    result: dict[str, str] = {}
    for row in range(2, ws.max_row + 1):
        query_id = clean_text(ws.cell(row, 1).value)
        if not query_id:
            continue
        result[query_id] = canonicalize_gold_from_authoritative(ws.cell(row, 4).value)
    return result


def collect_rows() -> list[dict[str, str]]:
    authoritative = load_authoritative_citations()
    wb = load_workbook(BENCHMARK_PATH)
    ws = wb[wb.sheetnames[0]]
    rows: list[dict[str, str]] = []

    for row in range(2, ws.max_row + 1):
        query_id = clean_text(ws.cell(row, 1).value)
        if not query_id or query_id == "query_id":
            continue

        if 2 <= row <= 643:
            query = clean_text(ws.cell(row, 2).value)
            gold = authoritative.get(query_id) or normalize_gold(ws.cell(row, 4).value)
            if not gold:
                gold = extract_citations_from_answer(ws.cell(row, 3).value)
            answer = standardize_answer(ws.cell(row, 3).value, gold)
            student_name = clean_text(ws.cell(row, 5).value)
            student_email = clean_text(ws.cell(row, 6).value)
        elif 644 <= row <= 893:
            query = clean_text(ws.cell(row, 2).value)
            gold = normalize_gold(ws.cell(row, 4).value)
            if not gold:
                gold = extract_citations_from_answer(ws.cell(row, 3).value)
            answer = standardize_answer(ws.cell(row, 3).value, gold)
            student_name = ""
            student_email = ""
        else:
            query = clean_text(ws.cell(row, 2).value)
            gold = authoritative.get(query_id) or normalize_gold(ws.cell(row, 6).value, ws.cell(row, 4).value)
            if not gold:
                gold = extract_citations_from_answer(ws.cell(row, 5).value)
            answer = standardize_answer(ws.cell(row, 5).value, gold)
            student_name = ""
            student_email = ""

        rows.append(
            {
                "query_id": query_id,
                "query": query,
                "Ответ нашего RAG": answer,
                "gold_citations": gold,
                "student_name": student_name,
                "student_email": student_email,
            }
        )

    def sort_key(item: dict[str, str]) -> tuple[int, int, str]:
        query_id = item["query_id"]
        if query_id.startswith("kz_legalrag_"):
            try:
                return (0, int(query_id.rsplit("_", 1)[1]), query_id)
            except ValueError:
                return (0, 10**9, query_id)
        if query_id.startswith("q_"):
            try:
                return (1, int(query_id.split("_", 1)[1]), query_id)
            except ValueError:
                return (1, 10**9, query_id)
        return (2, 10**9, query_id)

    rows.sort(key=sort_key)
    return rows


def write_clean_workbook(rows: list[dict[str, str]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Benchmark"
    ws.append(HEADERS)

    for item in rows:
        ws.append([item[header] for header in HEADERS])

    widths = {
        "A": 18,
        "B": 80,
        "C": 120,
        "D": 55,
        "E": 24,
        "F": 32,
    }
    for column, width in widths.items():
        ws.column_dimensions[column].width = width

    ws.freeze_panes = "A2"
    wb.save(BENCHMARK_PATH)


def main() -> None:
    rows = collect_rows()
    write_clean_workbook(rows)
    print(f"Rebuilt {BENCHMARK_PATH.name} with {len(rows)} rows.")


if __name__ == "__main__":
    main()

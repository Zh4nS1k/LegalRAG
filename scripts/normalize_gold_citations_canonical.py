from __future__ import annotations

import argparse
import re
from collections import OrderedDict, defaultdict
from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import load_workbook


BASE_DIR = Path(__file__).resolve().parents[1]
DEFAULT_INPUT = BASE_DIR / "tests" / "benchmarks" / "Полный бенчмарк-3.xlsx"
DEFAULT_OUTPUT = BASE_DIR / "tests" / "benchmarks" / "Полный бенчмарк-3.normalized.xlsx"
DEFAULT_SHEET = "Benchmark"

LAW_TITLE_RE = re.compile(r"закон[а-я\s]*(?:республики казахстан|рк)?\s*[«\"]([^»\"]+)[»\"]", re.IGNORECASE)
ORDER_RE = re.compile(r"(приказ[^.;\n]*?№\s*[\w/-]+[^.;\n]*)", re.IGNORECASE)
RESOLUTION_RE = re.compile(r"(постановлени[ея][^.;\n]*)", re.IGNORECASE)
EXPLICIT_ACT_PATTERNS: list[tuple[re.Pattern[str], str | None]] = [
    (re.compile(r"\bгк рк\b", re.IGNORECASE), "ГК РК"),
    (re.compile(r"\bгражданск[а-я\s]+кодекс[а-я\s]*(?:республики казахстан|рк)?\b", re.IGNORECASE), "ГК РК"),
    (re.compile(r"\bгпк рк\b", re.IGNORECASE), "ГПК РК"),
    (re.compile(r"\bгражданск[а-я\s]+процессуальн[а-я\s]+кодекс[а-я\s]*(?:республики казахстан|рк)?\b", re.IGNORECASE), "ГПК РК"),
    (re.compile(r"\bтк рк\b", re.IGNORECASE), "ТК РК"),
    (re.compile(r"\bтрудов[а-я\s]+кодекс[а-я\s]*(?:республики казахстан|рк)?\b", re.IGNORECASE), "ТК РК"),
    (re.compile(r"\bнк рк\b", re.IGNORECASE), "НК РК"),
    (re.compile(r"\bналогов[а-я\s]+кодекс[а-я\s]*(?:республики казахстан|рк)?\b", re.IGNORECASE), "НК РК"),
    (re.compile(r"\bкоап рк\b", re.IGNORECASE), "КоАП РК"),
    (re.compile(r"административных правонарушениях", re.IGNORECASE), "КоАП РК"),
    (re.compile(r"\bск рк\b", re.IGNORECASE), "Кодекс РК «О браке (супружестве) и семье»"),
    (re.compile(r"\bсемейн[а-я\s]+кодекс[а-я\s]*(?:республики казахстан|рк)?\b", re.IGNORECASE), "Кодекс РК «О браке (супружестве) и семье»"),
    (re.compile(r"\bкобс\b", re.IGNORECASE), "Кодекс РК «О браке (супружестве) и семье»"),
    (re.compile(r"о браке \(супружестве\) и семье", re.IGNORECASE), "Кодекс РК «О браке (супружестве) и семье»"),
    (re.compile(r"конституци[а-я\s]*(?:республики казахстан|рк)?", re.IGNORECASE), "Конституция РК"),
]

TOKEN_SPLIT_RE = re.compile(r"\s*;\s*")
SPACE_RE = re.compile(r"\s+")
POINT_RE = re.compile(r"\bп\.\s*(\d+(?:-\d+)?)", re.IGNORECASE)
PART_RE = re.compile(r"\bч\.\s*(\d+(?:-\d+)?)", re.IGNORECASE)
SUBPOINT_RE = re.compile(r"\bподп?\.\s*([0-9]+|[а-яa-z]+)\)?", re.IGNORECASE)
ARTICLE_RE = re.compile(r"\bст\.?\s*(\d+(?:[-–]\d+)?)", re.IGNORECASE)
@dataclass
class ParsedCitation:
    raw: str
    act: str = ""
    article: str = ""
    points: list[str] = field(default_factory=list)
    parts: list[str] = field(default_factory=list)
    subpoints: list[str] = field(default_factory=list)
    freeform_without_article: bool = False


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Normalize gold_citations into canonical legal format.")
    parser.add_argument("--input", default=str(DEFAULT_INPUT))
    parser.add_argument("--output", default=str(DEFAULT_OUTPUT))
    parser.add_argument("--sheet", default=DEFAULT_SHEET)
    return parser.parse_args()


def clean_text(value: object) -> str:
    if value is None:
        return ""
    text = str(value).replace("\xa0", " ").replace("\r\n", "\n").replace("\r", "\n").strip()
    text = SPACE_RE.sub(" ", text)
    text = re.sub(r"\s*;\s*", "; ", text)
    return text.strip(" ;")


def normalize_act_name(text: str) -> str:
    compact = clean_text(text)
    if not compact:
        return ""

    lowered = compact.lower()
    if "гк рк" in lowered or "гражданск" in lowered and "кодекс" in lowered:
        if "процессуал" not in lowered:
            return "ГК РК"
    if "гпк рк" in lowered or "гражданск" in lowered and "процессуал" in lowered and "кодекс" in lowered:
        return "ГПК РК"
    if "тк рк" in lowered or "трудов" in lowered and "кодекс" in lowered:
        return "ТК РК"
    if "нк рк" in lowered or "налогов" in lowered and "кодекс" in lowered:
        return "НК РК"
    if "коап рк" in lowered or "административных правонарушениях" in lowered:
        return "КоАП РК"
    if "ук рк" in lowered or "уголовного кодекса" in lowered:
        return "УК РК"
    if "упк рк" in lowered or "уголовно-процесс" in lowered:
        return "УПК РК"
    if "аппк рк" in lowered or "административного процедурно-процессуального кодекса" in lowered:
        return "АППК РК"
    if "пк рк" in lowered or "предпринимательского кодекса" in lowered:
        return "ПК РК"
    if "зк рк" in lowered or "земельного кодекса" in lowered:
        return "ЗК РК"
    if "конституц" in lowered:
        return "Конституция РК"
    if "кобс" in lowered or "о браке (супружестве) и семье" in lowered:
        return "Кодекс РК «О браке (супружестве) и семье»"

    law_match = LAW_TITLE_RE.search(compact)
    if law_match:
        return f"Закон РК «{law_match.group(1).strip()}»"

    order_match = ORDER_RE.search(compact)
    if order_match:
        return clean_text(order_match.group(1))

    resolution_match = RESOLUTION_RE.search(compact)
    if resolution_match:
        return clean_text(resolution_match.group(1))
    return ""


def normalize_num_list(items: list[str]) -> list[str]:
    seen: OrderedDict[str, None] = OrderedDict()
    for item in items:
        normalized = item.replace("–", "-").strip()
        if normalized:
            seen[normalized] = None
    return list(seen.keys())


def parse_citation(token: str) -> ParsedCitation:
    text = clean_text(token)
    parsed = ParsedCitation(raw=text)
    if not text:
        return parsed

    article_match = ARTICLE_RE.search(text)
    if article_match:
        parsed.article = article_match.group(1).replace("–", "-")
    parsed.points = normalize_num_list(POINT_RE.findall(text))
    parsed.parts = normalize_num_list(PART_RE.findall(text))
    parsed.subpoints = normalize_num_list(SUBPOINT_RE.findall(text))
    parsed.act = normalize_act_name(text)

    if not parsed.article and parsed.act:
        parsed.freeform_without_article = True
    if not parsed.article and (parsed.points or parsed.parts or parsed.subpoints):
        parsed.freeform_without_article = True
    if not parsed.article and not parsed.act and ("приказ" in text.lower() or "постанов" in text.lower()):
        parsed.freeform_without_article = True
    return parsed


def extract_context_citations(text: str) -> list[ParsedCitation]:
    cleaned = clean_text(text)
    if not cleaned:
        return []
    patterns = [
        re.compile(
            r"(?:(?:подп?\.\s*[0-9а-яa-z]+\)?\s*)?"
            r"(?:(?:п\.\s*\d+(?:-\d+)?\s*)|(?:ч\.\s*\d+(?:-\d+)?\s*))*)?"
            r"ст\.?\s*\d+(?:[-–]\d+)?[^.;\n]{0,180}?"
            r"(?:[А-Яа-яA-Za-z]+ кодекс[а-я\s]*|[Зз]акон[а-я\s]*[«\"][^»\"]+[»\"]|КоБС|ГК РК|ГПК РК|ТК РК|НК РК|КоАП РК|УК РК|УПК РК|АППК РК|ПК РК|ЗК РК)",
            re.IGNORECASE,
        ),
        re.compile(
            r"(?:[А-Яа-яA-Za-z]+ кодекс[а-я\s]*|[Зз]акон[а-я\s]*[«\"][^»\"]+[»\"]|КоБС|ГК РК|ГПК РК|ТК РК|НК РК|КоАП РК|УК РК|УПК РК|АППК РК|ПК РК|ЗК РК)"
            r"[^.;\n]{0,120}?ст\.?\s*\d+(?:[-–]\d+)?(?:[^.;\n]{0,60}?(?:п\.\s*\d+(?:-\d+)?|ч\.\s*\d+(?:-\d+)?|подп?\.\s*[0-9а-яa-z]+\)?))*",
            re.IGNORECASE,
        ),
    ]
    results: list[ParsedCitation] = []
    for pattern in patterns:
        results.extend(parse_citation(match.group(0)) for match in pattern.finditer(cleaned))
    dedup: OrderedDict[str, ParsedCitation] = OrderedDict()
    for item in results:
        if item.article and item.act:
            dedup[format_citation(item)] = item
    return list(dedup.values())


def extract_acts(text: str) -> list[str]:
    cleaned = clean_text(text)
    if not cleaned:
        return []
    acts: OrderedDict[str, None] = OrderedDict()
    for pattern, canonical in EXPLICIT_ACT_PATTERNS:
        for _ in pattern.finditer(cleaned):
            if canonical:
                acts[canonical] = None
    for law_match in LAW_TITLE_RE.finditer(cleaned):
        acts[f"Закон РК «{law_match.group(1).strip()}»"] = None
    for order_match in ORDER_RE.finditer(cleaned):
        acts[clean_text(order_match.group(1))] = None
    for resolution_match in RESOLUTION_RE.finditer(cleaned):
        acts[clean_text(resolution_match.group(1))] = None
    return list(acts.keys())


def infer_act(parsed: ParsedCitation, article_to_act: dict[str, str], row_acts: list[str], previous_act: str) -> str:
    if parsed.act:
        return parsed.act
    if parsed.article and parsed.article in article_to_act:
        return article_to_act[parsed.article]
    if previous_act:
        return previous_act
    if len(row_acts) == 1:
        return row_acts[0]
    return ""


def format_citation(parsed: ParsedCitation) -> str:
    if parsed.freeform_without_article:
        return parsed.raw

    chunks: list[str] = []
    if parsed.subpoints:
        prefix = "подп. " + ", ".join(parsed.subpoints)
        chunks.append(prefix)
    if parsed.points:
        chunks.append("п. " + ", ".join(parsed.points))
    if parsed.parts:
        chunks.append("ч. " + ", ".join(parsed.parts))
    if parsed.article:
        chunks.append(f"ст. {parsed.article}")
    if parsed.act:
        chunks.append(parsed.act)
    return " ".join(chunks).strip()


def normalize_row_gold(gold_text: object, answer_text: object) -> str:
    raw_tokens = [clean_text(part) for part in TOKEN_SPLIT_RE.split(clean_text(gold_text)) if clean_text(part)]
    parsed_gold = [parse_citation(token) for token in raw_tokens]
    parsed_context = [item for item in extract_context_citations(clean_text(answer_text)) if item.article and item.act]
    answer_acts = extract_acts(clean_text(answer_text))

    article_to_act: dict[str, str] = {}
    row_acts_ordered: OrderedDict[str, None] = OrderedDict()

    for item in parsed_context + parsed_gold:
        if item.act:
            row_acts_ordered[item.act] = None
        if item.article and item.act and item.article not in article_to_act:
            article_to_act[item.article] = item.act

    for act in answer_acts:
        row_acts_ordered[act] = None
    row_acts = list(row_acts_ordered.keys())

    row_articles = [item.article for item in parsed_gold if item.article]
    unique_row_articles = list(OrderedDict((article, None) for article in row_articles).keys())

    normalized_items: list[ParsedCitation] = []
    previous_act = ""
    for item in parsed_gold:
        if item.freeform_without_article:
            if not item.act:
                item.act = infer_act(item, article_to_act, row_acts, previous_act)
            if item.points and not item.article and len(unique_row_articles) == 1 and item.act:
                item.article = unique_row_articles[0]
                item.freeform_without_article = False
            normalized_items.append(item)
            previous_act = item.act or previous_act
            continue
        item.act = infer_act(item, article_to_act, row_acts, previous_act)
        if item.act:
            previous_act = item.act
        normalized_items.append(item)

    grouped: OrderedDict[tuple[str, str], dict[str, object]] = OrderedDict()
    passthrough: list[str] = []

    for item in normalized_items:
        if item.freeform_without_article:
            passthrough.append(item.raw)
            continue
        if not item.article or not item.act:
            passthrough.append(format_citation(item))
            continue

        key = (item.article, item.act)
        entry = grouped.setdefault(
            key,
            {
                "points": [],
                "parts": [],
                "subpoint_tokens": [],
                "has_article_only": False,
            },
        )

        if item.subpoints:
            entry["subpoint_tokens"].append(format_citation(item))
            continue
        if item.points:
            entry["points"].extend(item.points)
        if item.parts:
            entry["parts"].extend(item.parts)
        if not item.points and not item.parts:
            entry["has_article_only"] = True

    result: list[str] = []
    for (article, act), payload in grouped.items():
        points = normalize_num_list(payload["points"])
        parts = normalize_num_list(payload["parts"])
        subpoint_tokens = normalize_num_list(payload["subpoint_tokens"])
        has_article_only = bool(payload["has_article_only"])

        for token in subpoint_tokens:
            result.append(token)

        if points and parts:
            result.append(f"п. {', '.join(points)}; ч. {', '.join(parts)} ст. {article} {act}")
        elif points:
            result.append(f"п. {', '.join(points)} ст. {article} {act}")
        elif parts:
            result.append(f"ч. {', '.join(parts)} ст. {article} {act}")

        if has_article_only or (not points and not parts and not subpoint_tokens):
            result.append(f"ст. {article} {act}")

    result.extend(passthrough)
    result = [clean_text(item) for item in result if clean_text(item)]
    return "; ".join(OrderedDict.fromkeys(result))


def main() -> None:
    args = parse_args()
    input_path = Path(args.input)
    if not input_path.exists():
        raise FileNotFoundError(f"Workbook not found: {input_path}")

    workbook = load_workbook(input_path)
    if args.sheet not in workbook.sheetnames:
        raise ValueError(f"Worksheet not found: {args.sheet}")
    sheet = workbook[args.sheet]

    headers = [clean_text(value) for value in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
    gold_index = headers.index("gold_citations") + 1
    answer_index = headers.index("Ответ нашего RAG") + 1

    for row_number in range(2, sheet.max_row + 1):
        gold_value = sheet.cell(row_number, gold_index).value
        answer_value = sheet.cell(row_number, answer_index).value
        normalized = normalize_row_gold(gold_value, answer_value)
        sheet.cell(row_number, gold_index).value = normalized

    output_path = Path(args.output)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    print(f"Saved normalized workbook: {output_path}")


if __name__ == "__main__":
    main()

from __future__ import annotations

import argparse
import csv
import re
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook


ACT_MARKERS = (
    "гк",
    "гпк",
    "упк",
    "ук",
    "коап",
    "аппк",
    "нк",
    "тк",
    "пк",
    "зк",
    "кодекс",
    "закон",
    "конституц",
    "правил",
    "приказ",
    "постанов",
)

ARTICLE_MARKER_RE = re.compile(r"\b(ст\.?|статья|п\.|пункт|ч\.|часть|подп\.|подпункт)\b", re.IGNORECASE)
ARTICLE_NUMBER_RE = re.compile(r"\bст\.?\s*\d+(?:[-–]\d+)?\b", re.IGNORECASE)
MULTI_ARTICLE_RE = re.compile(r"\bст\.?\s*\d+(?:\s*,\s*\d+)+\b", re.IGNORECASE)
SPACE_RE = re.compile(r"\s+")


@dataclass(frozen=True)
class Finding:
    row_number: int
    query_id: str
    issue: str
    citation: str


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Audit gold citations in benchmark XLSX for legal-benchmark quality issues."
    )
    parser.add_argument(
        "--xlsx",
        default="tests/benchmarks/Полный бенчмарк-3.xlsx",
        help="Path to workbook with gold_citations column.",
    )
    parser.add_argument(
        "--sheet",
        default="Benchmark",
        help="Worksheet name.",
    )
    parser.add_argument(
        "--out",
        default="tests/benchmarks/gold_citations_audit.csv",
        help="CSV report path.",
    )
    return parser.parse_args()


def clean_text(value: object) -> str:
    if value is None:
        return ""
    text = str(value).replace("\xa0", " ").strip()
    return SPACE_RE.sub(" ", text)


def split_citations(value: object) -> list[str]:
    text = clean_text(value)
    if not text:
        return []
    return [part.strip(" ;") for part in text.split(";") if part.strip(" ;")]


def has_act_name(text: str) -> bool:
    lowered = text.lower()
    return any(marker in lowered for marker in ACT_MARKERS)


def audit_citation(row_number: int, query_id: str, citation: str) -> list[Finding]:
    findings: list[Finding] = []
    normalized = clean_text(citation)
    lowered = normalized.lower()

    if not normalized:
        findings.append(Finding(row_number, query_id, "empty_citation", citation))
        return findings

    if ARTICLE_MARKER_RE.search(normalized) and not has_act_name(normalized):
        findings.append(Finding(row_number, query_id, "missing_act_name", normalized))

    if MULTI_ARTICLE_RE.search(lowered):
        findings.append(Finding(row_number, query_id, "multiple_articles_in_one_token", normalized))

    if normalized.endswith("."):
        findings.append(Finding(row_number, query_id, "trailing_period", normalized))

    if "  " in citation:
        findings.append(Finding(row_number, query_id, "double_spaces", normalized))

    if "ст." in lowered and not ARTICLE_NUMBER_RE.search(lowered):
        findings.append(Finding(row_number, query_id, "article_marker_without_number", normalized))

    if ("подп." in lowered or "п." in lowered or "ч." in lowered) and "ст." not in lowered and not has_act_name(normalized):
        findings.append(Finding(row_number, query_id, "structural_unit_without_article_or_act", normalized))

    return findings


def main() -> None:
    args = parse_args()
    xlsx_path = Path(args.xlsx)
    if not xlsx_path.exists():
        raise FileNotFoundError(f"Workbook not found: {xlsx_path}")

    workbook = load_workbook(xlsx_path, read_only=True, data_only=True)
    if args.sheet not in workbook.sheetnames:
        raise ValueError(f"Worksheet not found: {args.sheet}")

    sheet = workbook[args.sheet]
    headers = [clean_text(value) for value in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
    try:
        query_id_index = headers.index("query_id")
        gold_index = headers.index("gold_citations")
    except ValueError as exc:
        raise ValueError("Required columns query_id and gold_citations were not found.") from exc

    findings: list[Finding] = []
    for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        query_id = clean_text(row[query_id_index])
        for citation in split_citations(row[gold_index]):
            findings.extend(audit_citation(row_number, query_id, citation))

    out_path = Path(args.out)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with out_path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.writer(handle)
        writer.writerow(["row_number", "query_id", "issue", "citation"])
        for item in findings:
            writer.writerow([item.row_number, item.query_id, item.issue, item.citation])

    issue_counts: dict[str, int] = {}
    for item in findings:
        issue_counts[item.issue] = issue_counts.get(item.issue, 0) + 1

    print(f"Saved audit report: {out_path}")
    print(f"Total findings: {len(findings)}")
    for issue, count in sorted(issue_counts.items(), key=lambda pair: (-pair[1], pair[0])):
        print(f"{issue}: {count}")


if __name__ == "__main__":
    main()

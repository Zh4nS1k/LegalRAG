from __future__ import annotations

import argparse
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook, load_workbook


DEFAULT_INPUT = "tests/benchmarks/kz_benchmark_gold_final.filtered.xlsx"
DEFAULT_OUTPUT = "tests/benchmarks/kz_benchmark_commercial_civil.filtered.xlsx"
DEFAULT_SHEET = "Benchmark"

POSITIVE_MARKERS = [
    "гражданского кодекса",
    "предпринимательского кодекса",
    "о товариществах с ограниченной и дополнительной ответственностью",
    "о реабилитации и банкротстве",
    "о банках и банковской деятельности",
    "о рынке ценных бумаг",
    "договор",
    "обязательств",
    "неустойк",
    "убытк",
    "поставка",
    "подряд",
    "аренд",
    "купл",
    "продаж",
    "кредит",
    "займ",
    "поручитель",
    "гарант",
    "тоо",
    "дивиденд",
    "корпоратив",
]

NEGATIVE_MARKERS = [
    "уголов",
    "административ",
    "трудов",
    "налогов",
    "семь",
    "браке",
    "социальн",
    "пенси",
    "земельн",
    "эколог",
    "миграц",
    "тамож",
    "пдд",
]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Filter benchmark rows to commercial-law focus (civil law) "
            "using deterministic keyword markers."
        )
    )
    parser.add_argument("--input", default=DEFAULT_INPUT, help="Input benchmark XLSX.")
    parser.add_argument("--output", default=DEFAULT_OUTPUT, help="Output filtered XLSX.")
    parser.add_argument("--sheet", default=DEFAULT_SHEET, help="Worksheet name.")
    parser.add_argument(
        "--max-rows",
        type=int,
        default=0,
        help="Optional cap on kept rows after filtering (0 means keep all).",
    )
    return parser.parse_args()


def clean_text(value: object) -> str:
    if value is None:
        return ""
    return " ".join(str(value).replace("\xa0", " ").split()).strip()


def marker_hits(text: str, markers: Iterable[str]) -> int:
    lowered = text.lower()
    return sum(1 for marker in markers if marker in lowered)


def keep_row(query: str, citations: str) -> tuple[bool, int, int]:
    combined = f"{query}\n{citations}".strip().lower()
    positive_hits = marker_hits(combined, POSITIVE_MARKERS)
    negative_hits = marker_hits(combined, NEGATIVE_MARKERS)

    if positive_hits == 0:
        return False, positive_hits, negative_hits
    if negative_hits == 0:
        return True, positive_hits, negative_hits
    return positive_hits > negative_hits, positive_hits, negative_hits


def get_header_index_map(headers: list[str]) -> dict[str, int]:
    index_map = {header: idx for idx, header in enumerate(headers)}
    if "query" not in index_map:
        raise ValueError("Column 'query' is required in the input sheet.")
    if "gold_citations" not in index_map:
        raise ValueError("Column 'gold_citations' is required in the input sheet.")
    return index_map


def main() -> None:
    args = parse_args()
    in_path = Path(args.input)
    out_path = Path(args.output)

    source_wb = load_workbook(in_path, read_only=True, data_only=False)
    source_ws = source_wb[args.sheet]

    out_wb = Workbook()
    out_ws = out_wb.active
    out_ws.title = args.sheet

    header_row = next(source_ws.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = [clean_text(value) for value in header_row]
    out_ws.append(list(header_row))
    index_map = get_header_index_map(headers)

    kept = 0
    removed = 0
    considered = 0

    for row in source_ws.iter_rows(min_row=2, values_only=True):
        considered += 1
        query = clean_text(row[index_map["query"]])
        citations = clean_text(row[index_map["gold_citations"]])
        keep, _, _ = keep_row(query, citations)
        if not keep:
            removed += 1
            continue
        if args.max_rows > 0 and kept >= args.max_rows:
            break
        out_ws.append(list(row))
        kept += 1

    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_wb.save(out_path)

    print(f"Saved commercial/civil filtered benchmark: {out_path}")
    print(f"Input rows considered: {considered}")
    print(f"Kept rows: {kept}")
    print(f"Removed rows: {removed}")
    if args.max_rows > 0:
        print(f"Row cap requested: {args.max_rows}")


if __name__ == "__main__":
    main()

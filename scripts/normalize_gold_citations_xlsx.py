from __future__ import annotations

import argparse
import json
import re
import time
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable
from urllib.parse import urljoin

import requests
from bs4 import BeautifulSoup
from bs4.element import Tag
from openpyxl import load_workbook
from openpyxl.styles import Font
import urllib3


BASE_DIR = Path(__file__).resolve().parents[1]
ADILET_ROOT = "https://adilet.zan.kz"
SEARCH_URL = f"{ADILET_ROOT}/rus/search/docs"
DOC_URL = f"{ADILET_ROOT}/rus/docs"
CACHE_PATH = BASE_DIR / "tests" / "benchmarks" / ".adilet_cache.json"
REQUEST_DELAY_SEC = 0.05
TIMEOUT_SEC = 30

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)


KNOWN_DOC_IDS: dict[str, str] = {
    "ГК РК": "K940001000_",
    "Гражданского кодекса РК": "K940001000_",
    "Гражданского кодекса Республики Казахстан": "K940001000_",
    "ГПК РК": "K1500000377",
    "Гражданского процессуального кодекса РК": "K1500000377",
    "ТК РК": "K1500000414",
    "Трудового кодекса РК": "K1500000414",
    "НК РК": "K1700000120",
    "Налогового кодекса РК": "K1700000120",
    "КоАП РК": "K1400000235",
    "Кодекса РК «Об административных правонарушениях»": "K1400000235",
    "УК РК": "K1400000226",
    "Уголовного кодекса РК": "K1400000226",
    "УПК РК": "K1400000231",
    "Уголовно-процессуального кодекса РК": "K1400000231",
    "АППК РК": "K2000000350",
    "Административного процедурно-процессуального кодекса РК": "K2000000350",
    "ПК РК": "K1500000375",
    "Предпринимательского кодекса РК": "K1500000375",
    "Предпринимательского Кодекса РК": "K1500000375",
    "Конституции РК": "K950001000_",
    "ЗК РК": "K030000442_",
    "Земельного кодекса РК": "K030000442_",
    "Земельного Кодекса РК": "K030000442_",
    "Кодекса РК «О браке (супружестве) и семье»": "K1100000518",
    "Закона РК «О защите прав потребителей»": "Z100000274_",
    "Закона РК «О правах потребителей»": "Z100000274_",
    "Закона РК «Об исполнительном производстве и статусе судебных исполнителей»": "Z100000261_",
    "Закона РК «Об исполнительном производстве»": "Z100000261_",
    "Закона РК «О жилищных отношениях»": "Z970000094_",
    "Закона РК «О жилищных отношения»": "Z970000094_",
    "Закона РК «О банках и банковской деятельности»": "Z950002444_",
    "Закона РК «О банках и банковской деятельности в Республике Казахстан»": "Z950002444_",
    "Закона РК «О банках и банковской деятельности Республике Казахстан»": "Z950002444_",
    "Закона РК «О банках и банковской деятельности в РК»": "Z950002444_",
    "Закона РК «О товариществах с ограниченной и дополнительной ответственностью»": "Z980000220_",
    "Закона РК «О ТОО»": "Z980000220_",
    "Закона РК «Об АО»": "Z030000415_",
    "Закона РК «О нотариате»": "Z970000155_",
    "Закона РК «О государственной регистрации прав на недвижимое имущество»": "Z070000310_",
    "Закона РК «О государственной регистрации юридических лиц и учетной регистрации филиалов и представительств»": "Z950002198_",
    "Закона РК «О воинской службе и статусе военнослужащих»": "Z1200000561",
    "Закона РК «О Воинской службе и статусе военнослужащих»": "Z1200000561",
    "Закона РК «О воинской службе»": "Z1200000561",
    "Закона РК «Об образовании»": "Z070000319_",
    "Закона РК «О рекламе»": "Z030000461_",
    "Закона РК «О дорожном движении»": "Z1400000194",
    "Закона РК «Об оценочной деятельности в Республике Казахстан»": "Z1800000133",
    "Закона РК «О кредитных бюро и формировании кредитных историй»": "Z040000573_",
    "Закона РК «О кредитных бюро и формировании кредитных историй в Республике Казахстан»": "Z040000573_",
    "Закона РК «О цифровых активах»": "Z2300000193",
    "Закона РК «О персональных данных и их защите»": "Z1300000094",
    "Закона РК «О государственных закупках»": "Z2400000106",
    "Закона РК «Об адвокатской деятельности и юридической помощи»": "Z1800000176",
    "Закона РК «О валютном регулировании и валютном контроле»": "Z1800000167",
    "Закона РК «О государственной службе Республики Казахстан»": "Z1500000416",
    "Закона РК «О государственной службе РК»": "Z1500000416",
    "Закона РК «О государственной службе»": "Z1500000416",
    "Закона РК «О правах ребенка»": "Z020000345_",
    "Закона РК «О правах ребенка в Республике Казахстан»": "Z020000345_",
    "Закона РК «О правах ребенка Республики Казахстан»": "Z020000345_",
    "Закона РК «О коллекторской деятельности»": "Z1700000062",
    "Закона РК «О микрофинансовой деятельности»": "Z1200000056",
    "Закона РК «О восстановлении платежеспособности и банкротстве граждан Республики Казахстан»": "Z2200000178",
    "Закона РК «О реабилитации и банкротстве»": "Z1400000176",
    "Закона РК «Об обязательном страховании гражданско-правовой ответственности владельцев транспортных средств»": "Z030000446_",
    "Закона РК «О страховании ответственности владельцев транспортных средств»": "Z030000446_",
    "Закона РК «О жилищный отношениях»": "Z970000094_",
    "Закона РК «Об исполнительном производстве и статусе частных судебных исполнителей»": "Z100000261_",
    "Закона РК «Об исполнительном производстве и статусе судебного исполнителя»": "Z100000261_",
    "Закона РК «О пенсионном обеспечении в РК»": "Z1300000105",
    "Закона РК «О пенсионном обеспечении»": "Z1300000105",
    "Закона РК «О Защите прав потребителей»": "Z100000274_",
    "Закона РК «Об архитектурной, градостроительной и строительной деятельности»": "Z010000242_",
    "Закона РК «о государственной службе РК»": "Z1500000416",
    "Закона РК «О республиканском бюджете на 2022-2024 годы»": "Z2100000077",
    "Закона РК «О республиканском бюджете на 2022 – 2024 годы»": "Z2100000077",
}

SPECIAL_DOC_BY_TITLE_AND_ARTICLE: dict[tuple[str, str], str] = {
    ("ГПК РК", "291"): "K990000411_",
    ("Гражданского процессуального кодекса РК", "291"): "K990000411_",
    ("Закона РК «О государственных закупках»", "47"): "Z1500000434",
}

DISPLAY_NORMALIZATION: dict[str, str] = {
    "Гражданского кодекса РК": "ГК РК",
    "Гражданского кодекса Республики Казахстан": "ГК РК",
    "Гражданского процессуального кодекса РК": "ГПК РК",
    "Гражданского Процессуального Кодекса РК": "ГПК РК",
    "Трудового кодекса РК": "ТК РК",
    "Налогового кодекса РК": "НК РК",
    "Кодекса РК «Об административных правонарушениях»": "КоАП РК",
    "Уголовного кодекса РК": "УК РК",
    "Уголовно-процессуального кодекса РК": "УПК РК",
    "Административного процедурно-процессуального кодекса РК": "АППК РК",
    "Предпринимательского кодекса РК": "ПК РК",
    "Предпринимательского Кодекса РК": "ПК РК",
    "Земельного кодекса РК": "ЗК РК",
    "Земельного Кодекса РК": "ЗК РК",
}

REFERENCE_PREFIX_RE = re.compile(
    r"^\s*(?:(?:пп?|ч)\.?\s*[\d\-–)]+\s*)*(?:ст\.?\s*[\d\-–]+\s*)?(?:(?:пп?|ч)\.?\s*[\d\-–)]+\s*)*",
    re.IGNORECASE,
)
ARTICLE_RE = re.compile(r"ст\.?\s*([\d]+(?:[\-–][\d]+)?)", re.IGNORECASE)
CITATION_SPLIT_RE = re.compile(r"\s*;\s*")
SPACE_RE = re.compile(r"\s+")
ARTICLE_ANCHOR_RE = re.compile(
    r'<a name="(z\d+)"></a>\s*</?[^>]*>\s*Статья\s+([0-9]+(?:-[0-9]+)?)',
    re.IGNORECASE,
)
HYPERLINK_FORMULA_RE = re.compile(
    r'^=HYPERLINK\("(?P<url>[^"]*)",\s*"(?P<label>(?:[^"]|"")*)"\)$',
    re.IGNORECASE,
)


@dataclass
class CitationLink:
    text: str
    url: str | None


class AdiletResolver:
    def __init__(self) -> None:
        self.session = requests.Session()
        self.session.headers.update(
            {"User-Agent": "Mozilla/5.0 (compatible; LegalRAG/1.0)"}
        )
        self.doc_cache, self.anchor_cache = self._load_cache()

    def _load_cache(self) -> tuple[dict[str, str], dict[str, dict[str, str]]]:
        if not CACHE_PATH.exists():
            return {}, {}
        payload = json.loads(CACHE_PATH.read_text(encoding="utf-8"))
        return payload.get("doc_cache", {}), payload.get("anchor_cache", {})

    def save_cache(self) -> None:
        CACHE_PATH.parent.mkdir(parents=True, exist_ok=True)
        payload = {
            "doc_cache": self.doc_cache,
            "anchor_cache": self.anchor_cache,
        }
        CACHE_PATH.write_text(
            json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8"
        )

    def resolve(self, citation: str) -> CitationLink:
        normalized = normalize_citation(citation)
        article_number = extract_article_number(normalized)
        if not article_number:
            return CitationLink(text=normalized, url=None)

        doc_id = self._resolve_doc_id(normalized)
        if not doc_id:
            return CitationLink(text=normalized, url=None)

        anchor = self._resolve_anchor(doc_id, article_number)
        if not anchor:
            return CitationLink(text=normalized, url=f"{DOC_URL}/{doc_id}")
        return CitationLink(text=normalized, url=f"{DOC_URL}/{doc_id}#{anchor}")

    def _resolve_doc_id(self, citation: str) -> str | None:
        article_number = extract_article_number(citation)
        if "ГК РК" in citation or "Гражданского кодекса РК" in citation or "Гражданского кодекса Республики Казахстан" in citation:
            primary_article = first_article_number(article_number)
            if primary_article is not None:
                return "K990000409_" if primary_article >= 406 else "K940001000_"

        title = extract_title_for_search(citation)
        article_key = article_number.replace("–", "-") if article_number else ""
        if title and article_key:
            primary_article = article_key.split("-", 1)[0]
            for special_title, special_article in SPECIAL_DOC_BY_TITLE_AND_ARTICLE:
                if special_title in citation and special_article == primary_article:
                    return SPECIAL_DOC_BY_TITLE_AND_ARTICLE[(special_title, special_article)]

        for key in sorted(KNOWN_DOC_IDS, key=len, reverse=True):
            if key in citation:
                return KNOWN_DOC_IDS[key]

        if not title:
            return None
        if title in self.doc_cache:
            return self.doc_cache[title] or None

        doc_id = self._search_doc_id(title)
        self.doc_cache[title] = doc_id or ""
        return doc_id

    def _search_doc_id(self, title: str) -> str | None:
        try:
            response = self.session.get(
                SEARCH_URL, params={"fulltext": title}, timeout=TIMEOUT_SEC, verify=False
            )
            response.raise_for_status()
        except requests.RequestException:
            return None
        time.sleep(REQUEST_DELAY_SEC)

        soup = BeautifulSoup(response.text, "html.parser")
        candidates: list[tuple[str, str]] = []
        for a in soup.find_all("a", href=True):
            href = a["href"]
            if "/rus/docs/" not in href:
                continue
            text = clean_text(a.get_text(" ", strip=True))
            if not text:
                continue
            candidates.append((href, text))
        if not candidates:
            return None

        best_href = pick_best_candidate(title, candidates)
        if not best_href:
            return None
        match = re.search(r"/rus/docs/([^/#?]+)", best_href)
        return match.group(1) if match else None

    def _resolve_anchor(self, doc_id: str, article_number: str) -> str | None:
        article_key = article_number.replace("–", "-")
        doc_anchors = self.anchor_cache.get(doc_id)
        if not doc_anchors:
            doc_anchors = self._fetch_doc_anchors(doc_id)
            self.anchor_cache[doc_id] = doc_anchors
        anchor = doc_anchors.get(article_key)
        if not anchor and "-" in article_key:
            anchor = doc_anchors.get(article_key.split("-", 1)[0])
        if not anchor:
            doc_anchors = self._fetch_doc_anchors(doc_id)
            self.anchor_cache[doc_id] = doc_anchors
            anchor = doc_anchors.get(article_key)
            if not anchor and "-" in article_key:
                anchor = doc_anchors.get(article_key.split("-", 1)[0])
        return anchor

    def _fetch_doc_anchors(self, doc_id: str) -> dict[str, str]:
        try:
            response = self.session.get(
                f"{DOC_URL}/{doc_id}", timeout=TIMEOUT_SEC, verify=False
            )
            response.raise_for_status()
        except requests.RequestException:
            return {}
        time.sleep(REQUEST_DELAY_SEC)

        soup = BeautifulSoup(response.text, "html.parser")
        anchors: dict[str, str] = {}

        for tag in soup.find_all(re.compile(r"^h[1-6]$")):
            tag_id = str(tag.get("id") or "")
            text = clean_text(tag.get_text(" ", strip=True))
            match = re.search(r"Статья\s+([0-9]+(?:-[0-9]+)?)\b", text)
            if match and tag_id.startswith("z"):
                anchors.setdefault(match.group(1).replace("–", "-"), tag_id)
            elif match:
                next_anchor = find_next_anchor_after_heading(tag)
                if next_anchor:
                    anchors.setdefault(match.group(1).replace("–", "-"), next_anchor)

        for tag in soup.find_all(True):
            tag_id = str(tag.get("id") or "")
            text = clean_text(tag.get_text(" ", strip=True))
            match = re.search(r"Статья\s+([0-9]+(?:-[0-9]+)?)\b", text)
            if match and tag_id.startswith("z"):
                anchors.setdefault(match.group(1).replace("–", "-"), tag_id)
            elif match and tag.name in {"p", "b", "font"}:
                next_anchor = find_next_anchor_after_heading(tag)
                if next_anchor:
                    anchors.setdefault(match.group(1).replace("–", "-"), next_anchor)

        for tag in soup.find_all("a", attrs={"name": True}):
            anchor = str(tag.get("name") or "")
            if not anchor.startswith("z"):
                continue
            parent = tag.parent
            text = clean_text(parent.get_text(" ", strip=True) if parent else "")
            match = re.match(r"Статья\s+([0-9]+(?:-[0-9]+)?)\b", text)
            if match:
                anchors.setdefault(match.group(1).replace("–", "-"), anchor)

        if not anchors:
            for anchor, article in ARTICLE_ANCHOR_RE.findall(response.text):
                anchors.setdefault(article.replace("–", "-"), anchor)
        return anchors


def clean_text(value: str) -> str:
    value = unwrap_hyperlink_formula(value)
    value = value.replace("\xa0", " ")
    value = value.replace("« ", "«").replace(" »", "»")
    return SPACE_RE.sub(" ", value).strip()


def normalize_citation(value: str) -> str:
    text = clean_text(str(value or ""))
    text = re.sub(r"\bстатья\b", "ст.", text, flags=re.IGNORECASE)
    text = re.sub(r"(?<!\w)ст\.?\s*(?=\d)", "ст. ", text, flags=re.IGNORECASE)
    text = re.sub(r"(?<!\w)пп\.?\s*(?=[\d(])", "пп. ", text, flags=re.IGNORECASE)
    text = re.sub(r"(?<!\w)п\.?\s*(?=[\d(])", "п. ", text, flags=re.IGNORECASE)
    text = re.sub(r"(?<!\w)ч\.?\s*(?=\d)", "ч. ", text, flags=re.IGNORECASE)
    text = re.sub(r"\s*,\s*", ", ", text)
    text = re.sub(r"\s*;\s*", "; ", text)
    text = re.sub(r"\s*([–-])\s*", r"\1", text)
    for src, dst in sorted(DISPLAY_NORMALIZATION.items(), key=lambda item: len(item[0]), reverse=True):
        text = text.replace(src, dst)
    text = text.replace('Закона РК «о государственной службе РК»', 'Закона РК «О государственной службе РК»')
    text = text.replace('Закона РК «О Защите прав потребителей»', 'Закона РК «О защите прав потребителей»')
    return clean_text(text)


def extract_article_number(citation: str) -> str | None:
    match = ARTICLE_RE.search(citation)
    if not match:
        return None
    article = match.group(1).replace("–", "-")
    if "-" in article:
        return article.split("-", 1)[0] if article.count("-") > 1 else article
    return article


def first_article_number(article_number: str | None) -> int | None:
    if not article_number:
        return None
    match = re.match(r"(\d+)", article_number)
    return int(match.group(1)) if match else None


def find_next_anchor_after_heading(tag) -> str | None:
    for sibling in tag.next_siblings:
        name = getattr(sibling, "name", None)
        if name and re.match(r"^h[1-6]$", name):
            break
        sibling_id = str(getattr(sibling, "attrs", {}).get("id") or "")
        if sibling_id.startswith("z"):
            return sibling_id
        if isinstance(sibling, Tag):
            anchor_tag = sibling.find("a", attrs={"name": True})
            if anchor_tag:
                anchor_name = str(anchor_tag.get("name") or "")
                if anchor_name.startswith("z"):
                    return anchor_name
    return None


def extract_title_for_search(citation: str) -> str | None:
    patterns = [
        r"(Закона РК «[^»]+»)",
        r"(Кодекса РК «[^»]+»)",
        r"(Гражданского кодекса РК)",
        r"(Гражданского кодекса Республики Казахстан)",
        r"(Гражданского процессуального кодекса РК)",
        r"(Трудового кодекса РК)",
        r"(Уголовного кодекса РК)",
        r"(Уголовно-процессуального кодекса РК)",
        r"(Налогового кодекса РК)",
        r"(Конституции РК)",
        r"(Предпринимательского [Кк]одекса РК)",
        r"(Административного процедурно-процессуального кодекса РК)",
        r"(Земельного [Кк]одекса РК)",
    ]
    for pattern in patterns:
        match = re.search(pattern, citation)
        if match:
            return match.group(1)

    stripped = REFERENCE_PREFIX_RE.sub("", citation).strip(" ,.")
    return stripped or None


def normalize_search_key(value: str) -> str:
    value = clean_text(value).lower()
    replacements = {
        "республики казахстан": "рк",
        "республики казахстана": "рк",
        "закона рк": "",
        "кодекса рк": "",
        "закон рк": "",
        "кодекс рк": "",
        "«": "",
        "»": "",
        '"': "",
    }
    for src, dst in replacements.items():
        value = value.replace(src, dst)
    value = re.sub(r"[^a-zа-яё0-9]+", " ", value, flags=re.IGNORECASE)
    return clean_text(value)


def pick_best_candidate(title: str, candidates: Iterable[tuple[str, str]]) -> str | None:
    normalized_title = normalize_search_key(title)
    quoted_match = re.search(r"«([^»]+)»", title)
    quoted = normalize_search_key(quoted_match.group(1)) if quoted_match else ""

    scored: list[tuple[int, int, str]] = []
    for href, text in candidates:
        normalized_text = normalize_search_key(text)
        score = 0
        if normalized_text == normalized_title:
            score += 100
        if quoted and normalized_text == quoted:
            score += 95
        if quoted and quoted in normalized_text:
            score += 60
        if normalized_title and normalized_title in normalized_text:
            score += 40
        if normalized_text and normalized_text in normalized_title:
            score += 25
        if "о мерах по реализации" in normalized_text:
            score -= 40
        if "об утверждении" in normalized_text:
            score -= 25
        scored.append((score, -len(text), href))

    scored.sort(reverse=True)
    best_score, _, best_href = scored[0]
    return best_href if best_score > 0 else None


def split_citations(value: str) -> list[str]:
    return [part for part in CITATION_SPLIT_RE.split(str(value or "")) if clean_text(part)]


def unwrap_hyperlink_formula(value: str) -> str:
    match = HYPERLINK_FORMULA_RE.match(str(value or "").strip())
    if not match:
        return str(value or "")
    return match.group("label").replace('""', '"')


def update_workbook(path: Path, backup: bool) -> None:
    if backup:
        backup_path = path.with_name(f"{path.stem}.backup{path.suffix}")
        if not backup_path.exists():
            backup_path.write_bytes(path.read_bytes())

    workbook = load_workbook(path)
    worksheet = workbook[workbook.sheetnames[0]]

    headers = [worksheet.cell(1, col).value for col in range(1, worksheet.max_column + 1)]
    try:
        gold_col_idx = headers.index("gold_citations") + 1
    except ValueError as exc:
        raise ValueError("В XLSX не найден столбец gold_citations") from exc

    resolver = AdiletResolver()
    rows_payload: list[tuple[int, list[CitationLink], str]] = []
    max_links = 0

    for row_idx in range(2, worksheet.max_row + 1):
        raw_value = worksheet.cell(row_idx, gold_col_idx).value
        parts = split_citations(str(raw_value or ""))
        links = []
        seen_pairs: set[tuple[str, str | None]] = set()
        for part in parts:
            link = resolver.resolve(part)
            key = (link.text, link.url)
            if key in seen_pairs:
                continue
            seen_pairs.add(key)
            links.append(link)
        normalized_cell = "; ".join(item.text for item in links)
        rows_payload.append((row_idx, links, normalized_cell))
        max_links = max(max_links, len(links))
        if row_idx % 100 == 0:
            print(f"processed {row_idx - 1} rows")

    while worksheet.max_column > gold_col_idx:
        worksheet.delete_cols(gold_col_idx + 1)

    for offset in range(max_links):
        title_cell = worksheet.cell(1, worksheet.max_column + 1)
        title_cell.value = f"citation_{offset + 1}"
        title_cell.font = Font(bold=True)
        url_cell = worksheet.cell(1, worksheet.max_column + 1)
        url_cell.value = f"adilet_url_{offset + 1}"
        url_cell.font = Font(bold=True)

    hyperlink_font = Font(color="0563C1", underline="single")

    for row_idx, links, normalized_cell in rows_payload:
        gold_cell = worksheet.cell(row_idx, gold_col_idx)
        gold_cell.value = normalized_cell
        gold_cell.hyperlink = None
        gold_cell.style = "Normal"

        gold_cell.value = normalized_cell

        for offset in range(max_links):
            citation_cell = worksheet.cell(row_idx, gold_col_idx + 1 + offset * 2)
            url_cell = worksheet.cell(row_idx, gold_col_idx + 2 + offset * 2)
            citation_cell.value = None
            citation_cell.hyperlink = None
            citation_cell.style = "Normal"
            url_cell.value = None
            url_cell.hyperlink = None
            url_cell.style = "Normal"
            if offset >= len(links):
                continue
            link = links[offset]
            citation_cell.value = link.text
            if link.url:
                url_cell.value = link.url
                url_cell.hyperlink = link.url
                url_cell.font = hyperlink_font

    worksheet.freeze_panes = "A2"
    worksheet.column_dimensions[worksheet.cell(1, gold_col_idx).column_letter].width = 55
    for offset in range(max_links):
        citation_letter = worksheet.cell(1, gold_col_idx + 1 + offset * 2).column_letter
        url_letter = worksheet.cell(1, gold_col_idx + 2 + offset * 2).column_letter
        worksheet.column_dimensions[citation_letter].width = 38
        worksheet.column_dimensions[url_letter].width = 55

    resolver.save_cache()
    workbook.save(path)
    print(f"saved {path}")


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--xlsx",
        default="tests/benchmarks/642_questions_with_citations.xlsx",
        help="Путь к XLSX-файлу.",
    )
    parser.add_argument(
        "--no-backup",
        action="store_true",
        help="Не создавать backup-копию файла перед изменением.",
    )
    args = parser.parse_args()

    update_workbook(Path(args.xlsx), backup=not args.no_backup)


if __name__ == "__main__":
    main()
